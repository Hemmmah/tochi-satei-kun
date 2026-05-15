"""2シート構成の xlsx 生成。
業者用シート（係数全開示）+ 顧客用シート（流推方式準拠、禁止語チェック）。

確定方針（プラン §7）：
- 業者用：n, R², 全β、補正内訳、参考値、警告
- 顧客用：ですます調、禁止語ブロック
"""
import math
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList

# グラフ用の短縮ラベル（表は詳細ラベル、グラフだけ短縮）
SHORT_FEATURE_LABELS = {
    "ln_area": "面積",
    "ln_area_sq": "面積²",
    "walk_min": "駅徒歩",
    "ln_shape": "形状指数",
    "ln_road_w": "道路幅員",
    "ln_far": "容積率",
    "dir_score": "方位",
    "D_shidou": "私道",
    "D_fukuro": "袋地",
    "D_fuseikei": "不整形",
    "ln_district_mean": "地区平均",
    "ln_station_mean": "駅勢圏",
    "const": "定数項",
}

from forbidden_words import assert_clean

# ===== スタイル =====
TITLE_FONT = Font(name="游ゴシック", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="2F5496")
SECTION_FONT = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="4472C4")
LABEL_FONT = Font(name="游ゴシック", size=10, bold=True)
VALUE_FONT = Font(name="游ゴシック", size=10)
BIG_VALUE_FONT = Font(name="游ゴシック", size=18, bold=True, color="C00000")
WARN_FILL = PatternFill("solid", fgColor="FFE699")
MISSING_FILL = PatternFill("solid", fgColor="F4B084")
P_LOW_FILL = PatternFill("solid", fgColor="C6E0B4")  # p<0.05
P_MID_FILL = PatternFill("solid", fgColor="FFE699")  # 0.05<=p<0.1
P_HIGH_FILL = PatternFill("solid", fgColor="F4B084")  # p>=0.1
PRIMARY_FILL = PatternFill("solid", fgColor="E2EFDA")  # 規範性の高い事例の薄緑（モジュール共通）
PRIMARY_FONT = Font(name="游ゴシック", size=10, bold=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _format_jpy(v):
    if v is None:
        return ""
    return f"{int(round(v)):,}円"


_CITY_CODE_TO_SHORT = {
    "13101": "千代田", "13102": "中央", "13103": "港", "13104": "新宿",
    "13105": "文京", "13106": "台東", "13107": "墨田", "13108": "江東",
    "13109": "品川", "13110": "目黒", "13111": "大田", "13112": "世田谷",
    "13113": "渋谷", "13114": "中野", "13115": "杉並", "13116": "豊島",
    "13117": "北", "13118": "荒川", "13119": "板橋", "13120": "練馬",
    "13121": "足立", "13122": "葛飾", "13123": "江戸川",
}


def _short_koji_id(std_id: str) -> str:
    """公示番号を短縮表記。例: '13112-000-050' → '世田谷-50'。"""
    if not std_id:
        return ""
    parts = str(std_id).split("-")
    if len(parts) < 3:
        return str(std_id)
    city_code = parts[0]
    point_num = parts[2]
    city_short = _CITY_CODE_TO_SHORT.get(city_code, city_code)
    try:
        n = int(point_num)
    except (TypeError, ValueError):
        n = point_num
    return f"{city_short}-{n}"


def _koji_shape_label(frontage_ratio, depth_ratio) -> str:
    """公示地点の間口比率(L01_036)・奥行比率(L01_037)から形状ラベルを推定。
    ratio ≤ 1.5 → 整形、≤ 2.5 → やや細長、それ以上 → 細長/不整形。
    """
    if frontage_ratio in (None, "", "_") or depth_ratio in (None, "", "_"):
        return "—"
    try:
        f = float(frontage_ratio)
        d = float(depth_ratio)
    except (TypeError, ValueError):
        return "—"
    if f <= 0 or d <= 0:
        return "—"
    ratio = max(d / f, f / d)
    if ratio <= 1.5:
        return "整形"
    if ratio <= 2.5:
        return "やや細長"
    return "細長"


def _short_koji_addr(address: str, district: str) -> str:
    """公示地点の所在を「地区名+丁目数字」に短縮。
    例: "東京都　世田谷区赤堤５丁目４８４番４" → "赤堤５"
    丁目がない住所は district を返す。
    """
    if not address:
        return district or ""
    if district and district in address:
        after = address[address.index(district):]
        if "丁目" in after:
            return after.split("丁目")[0]
        return district
    # district が住所に見つからない場合は丁目までを返す
    if "丁目" in address:
        # 都道府県・市区町村を除去（最後の "区"/"市"/"町"/"村" 以降）
        for sep in ["区", "市", "町", "村"]:
            if sep in address:
                idx = address.rindex(sep)
                addr_local = address[idx+1:]
                if "丁目" in addr_local:
                    return addr_local.split("丁目")[0]
        return address.split("丁目")[0]
    return district or address


def _format_pct(v):
    if v is None:
        return ""
    return f"{v*100:+.2f}%"


def _format_hijun_corr(multiplier, applies=True, mode="auto"):
    """比準表用の補正値表記（顧客用シート、1セル文字列）。
    mode:
      "top"    = 常に分子側（時点修正：査定時点 / 事例時点）
      "bottom" = 常に分母側（標準化補正・地域格差：100 / 事例評点）
      "auto"   = 補正方向で自動切替
    """
    if not applies:
        return "100/-"
    if abs(multiplier - 1.0) < 0.0005:
        return "100/100"
    if mode == "top":
        return f"{multiplier*100:.1f}/100"
    if mode == "bottom":
        # 上=100, 下=案件評点(=mult*100) — 倍率 = 下/上
        return f"100/{multiplier*100:.1f}"
    if multiplier > 1.0:
        return f"{multiplier*100:.1f}/100"
    return f"100/{multiplier*100:.1f}"


def _hijun_top_bottom(multiplier, applies=True, mode="auto"):
    """比準表用の分子/分母を別々に返す（業者用シートの2行式表示）。
    mode は _format_hijun_corr と同じ。
    Returns: (top, bottom) — 数値または "―"
    """
    if not applies:
        return (100, "―")
    if abs(multiplier - 1.0) < 0.0005:
        return (100, 100)
    if mode == "top":
        return (round(100 * multiplier, 1), 100)
    if mode == "bottom":
        # 上=100, 下=案件評点(=mult*100) — 倍率 = 下/上
        return (100, round(100 * multiplier, 1))
    if multiplier > 1.0:
        return (round(100 * multiplier, 1), 100)
    return (100, round(100 * multiplier, 1))


def _round_3sig(n):
    """上位3桁に四捨五入。例: 424,674,476 → 425,000,000"""
    if n is None or n == 0:
        return n
    import math
    sign = -1 if n < 0 else 1
    n = abs(n)
    digits = int(math.log10(n)) + 1
    if digits <= 3:
        return sign * int(round(n))
    factor = 10 ** (digits - 3)
    return sign * int(round(n / factor) * factor)


def _format_price_full(total_price, area):
    """査定価格表記: 「総額（〇円/㎡、〇円/坪）」を上位3桁四捨五入で。
    坪単価 = ㎡単価 ÷ 0.3025
    """
    if total_price is None or area is None or area <= 0:
        return ""
    total_r = _round_3sig(total_price)
    unit_per_sqm = total_r / area
    unit_per_sqm_r = _round_3sig(unit_per_sqm)
    unit_per_tsubo_r = _round_3sig(unit_per_sqm / 0.3025)
    return f"{total_r:,}円（{unit_per_sqm_r:,}円/㎡、{unit_per_tsubo_r:,}円/坪）"


def _set(ws, row, col, value, font=None, fill=None, align=None, border=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = BORDER
    if number_format: c.number_format = number_format
    return c


def _section_header(ws, row, text, end_col=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    _set(ws, row, 1, text, font=SECTION_FONT, fill=SECTION_FILL,
         align=Alignment(horizontal="left", vertical="center"))


def _adjust_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ===== 業者用シート =====
def _write_gyosha_sheet(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("業者用")
    # グラフ専用シートを 業者用 の直後（インデックス 1）に作成
    graph_ws = wb.create_sheet("グラフ", 1)
    # グラフシートのタイトル
    _set(graph_ws, 1, 1, "■ 附属資料",
         font=Font(name="游ゴシック", size=14, bold=True, color="FFFFFF"),
         fill=TITLE_FILL,
         align=Alignment(horizontal="left", vertical="center"))
    graph_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    graph_ws.row_dimensions[1].height = 28
    # 列幅をグラフ表示用に調整
    for col_letter in 'ABCDEFGHIJKLMN':
        graph_ws.column_dimensions[col_letter].width = 10
    # グラフ配置用の running row tracker
    ctx['_graph_ws'] = graph_ws
    ctx['_graph_row'] = 3  # タイトル(1)+空行(2)
    target = ctx["target"]
    asof = ctx["asof"]
    scope_log = ctx["scope_log"]
    rate_info = ctx["rate_info"]
    hed = ctx["hedonic"]
    cases = ctx["cases"]
    breakdown = ctx["breakdown"]
    assess = ctx["assess"]
    refs = ctx["refs"]
    standard_check = ctx["standard_check"]
    hijun_rows = ctx.get("hijun_rows", [])
    hijun_detail_rows = ctx.get("hijun_detail_rows", [])

    r = 1
    # タイトル
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(ws, r, 1, f"土地価格査定 業者用シート — {target.get('物件略号', '')} ({target['市区町村名']} {target.get('地区名', '')})",
         font=TITLE_FONT, fill=TITLE_FILL,
         align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[r].height = 28
    r += 2

    # ヘッダ：物件概要
    _section_header(ws, r, "■ 物件概要・スコープ")
    r += 1
    info = [
        ("査定時点", asof.isoformat()),
        ("所在", f"{target['都道府県名']} {target['市区町村名']} {target.get('地区名', '')}{target.get('丁目', '')}"),
        ("面積", f"{target['面積(㎡)']} ㎡"),
        ("最寄駅", f"{target.get('最寄駅:名称', '')} 徒歩{target.get('最寄駅:距離(分)', '')}分"),
        ("形状", target.get("土地の形状", "")),
        ("接道", f"{target.get('前面道路:種類', '')} 幅員{target.get('前面道路:幅員(m)', '')}m {target.get('前面道路:方位', '')}向"),
        ("用途地域", target.get("都市計画", "")),
        ("建ぺい率/容積率", f"{target.get('建ぺい率(%)', '')}% / {target.get('容積率(%)', '')}%"),
        ("使用事例件数", f"{scope_log['final_count']} 件 (IQR除外: {scope_log['iqr_removed']}件、市区町村単位・隣接拡張なし)"),
    ]
    for label, value in info:
        _set(ws, r, 1, label, font=LABEL_FONT, border=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        _set(ws, r, 2, value, font=VALUE_FONT, border=True)
        r += 1
    r += 1

    # 査定価格
    _section_header(ws, r, "■ 査定価格")
    r += 1
    target_area = target["面積(㎡)"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(ws, r, 1,
         _format_price_full(assess["central_total_price"], target_area),
         font=BIG_VALUE_FONT,
         align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[r].height = 36
    r += 1

    # 信頼度ラベル（高/中/中-低/低）：n × 自由度調整済 R² × 期待符号整合性ベース
    # 中-低 は「構造問題」（**統計的に有意な**符号反転2件以上 or adj_R² が極端に低い）に限定。
    # 非有意（p≥0.10）な符号反転はノイズ範囲内とみなしカウントしない。
    if hed["ok"]:
        n = hed["n"]
        adj_r2 = hed["adj_r2"]
        EXPECTED_NEG = ("ln_area", "walk_min", "D_shidou", "D_fukuro", "D_fuseikei")
        SIG_P_THRESHOLD = 0.10  # この p 値未満の反転のみ「有意な反転」としてカウント
        coef = hed["coefficients"]
        sign_inconsistent = sum(
            1 for name in EXPECTED_NEG
            if name in coef and coef[name]["beta"] > 0 and coef[name]["p"] < SIG_P_THRESHOLD
        )
        sign_checked = sum(1 for name in EXPECTED_NEG if name in coef)
        if sign_inconsistent >= 2 or adj_r2 < 0.3:
            reasons = []
            if sign_inconsistent >= 2:
                reasons.append(f"有意な符号反転 {sign_inconsistent}/{sign_checked} 件（p<{SIG_P_THRESHOLD}）")
            if adj_r2 < 0.3:
                reasons.append(f"adj R² = {adj_r2:.2f}（低水準）")
            conf_label = (f"信頼度：中-低（n = {n}, "
                          + ", ".join(reasons)
                          + " — 構造問題の可能性、要再確認）")
            conf_fill = P_HIGH_FILL
        elif n >= 20 and adj_r2 >= 0.45 and sign_inconsistent == 0:
            conf_label = (f"信頼度：高（n = {n}, 自由度調整済 R² = {adj_r2:.2f}, "
                          f"有意な期待符号と全整合）")
            conf_fill = P_LOW_FILL
        else:
            reasons = []
            if n < 20:
                reasons.append(f"事例件数 n = {n} と少なめ")
            if adj_r2 < 0.45:
                reasons.append(f"adj R² = {adj_r2:.2f}（中程度）")
            if sign_inconsistent == 1:
                reasons.append(f"有意な符号反転 1/{sign_checked} 件")
            if not reasons:
                reasons.append(f"n = {n}, adj R² = {adj_r2:.2f}")
            conf_label = "信頼度：中（" + " / ".join(reasons) + "）"
            conf_fill = P_MID_FILL
    else:
        conf_label = "信頼度：低（件数不足のため係数推定不能。顧客用シートは『参考情報』として出力）"
        conf_fill = P_HIGH_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(ws, r, 1, conf_label,
         font=Font(name="游ゴシック", size=11, bold=True),
         fill=conf_fill, border=True,
         align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[r].height = 24
    r += 2

    # 2価格サマリ：採用査定価格 vs ヘドニック母集団予測（信頼度ラベル直下に配置）
    target_area_local = target["面積(㎡)"]
    central_unit = assess.get("central_unit_price")
    hed_pred = refs.get("hedonic_pred")

    _section_header(ws, r, "■ 2価格サマリ（採用査定価格 vs ヘドニック母集団予測の乖離）")
    r += 1
    for j, h in enumerate(["区分", "㎡単価", "総額", "採用査定との乖離率"]):
        _set(ws, r, j+1, h, font=LABEL_FONT,
             fill=PatternFill("solid", fgColor="D9E1F2"), border=True)
    r += 1

    # ① 土地比準（採用）
    _set(ws, r, 1, "① 土地比準（採用）",
         font=Font(name="游ゴシック", size=10, bold=True),
         fill=PRIMARY_FILL, border=True)
    if central_unit:
        _set(ws, r, 2, f"{int(central_unit):,}円",
             font=Font(name="游ゴシック", size=10, bold=True), fill=PRIMARY_FILL, border=True)
        _set(ws, r, 3, _format_jpy(central_unit * target_area_local),
             font=Font(name="游ゴシック", size=10, bold=True), fill=PRIMARY_FILL, border=True)
    _set(ws, r, 4, "—", font=VALUE_FONT, fill=PRIMARY_FILL, border=True)
    r += 1

    # ② ヘドニック母集団予測
    if hed_pred and central_unit:
        dev = (hed_pred - central_unit) / central_unit * 100
        abs_dev = abs(dev)
        if abs_dev <= 15:
            dev_fill = P_LOW_FILL
            dev_guide = "（15%以内：採用査定とヘドニック予測が概ね整合）"
        elif abs_dev <= 30:
            dev_fill = P_MID_FILL
            dev_guide = "（15〜30%：地域特性または個別事例の特殊性を確認すると良い）"
        else:
            dev_fill = P_HIGH_FILL
            dev_guide = ("※ 30%超：規範性の高い事例が母集団から外れている可能性。"
                         "事例選定と特徴量を再確認してください。")
        _set(ws, r, 1, "② ヘドニック母集団予測", font=VALUE_FONT, border=True)
        _set(ws, r, 2, f"{int(hed_pred):,}円", font=VALUE_FONT, border=True)
        _set(ws, r, 3, _format_jpy(hed_pred * target_area_local), font=VALUE_FONT, border=True)
        _set(ws, r, 4, f"{dev:+.1f}%", font=VALUE_FONT, fill=dev_fill, border=True)
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        _set(ws, r, 1, dev_guide,
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"))
        r += 1
    elif central_unit:
        _set(ws, r, 1, "② ヘドニック母集団予測", font=VALUE_FONT, border=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        _set(ws, r, 2, "（件数不足のため算出不能）", font=VALUE_FONT, fill=MISSING_FILL, border=True)
        r += 1
    r += 1

    # 価格レンジ（比準表の試算値の最大/中央/最小と一致）
    _section_header(ws, r, "■ 価格レンジ（比準表の試算値 最大／中央／最小）")
    r += 1
    rng = assess["range"]
    headers = ["区分", "総額", "㎡単価", "坪単価"]
    for j, h in enumerate(headers):
        _set(ws, r, j+1, h, font=LABEL_FONT, fill=PatternFill("solid", fgColor="D9E1F2"), border=True)
    r += 1
    for label, total, unit in [
        ("上限", rng["high_total"], rng["high_unit"]),
        ("中央", rng["central_total"], rng["central_unit"]),
        ("下限", rng["low_total"], rng["low_unit"]),
    ]:
        total_r = _round_3sig(total) if total else None
        unit_sqm_r = _round_3sig(unit) if unit else None
        unit_tsubo_r = _round_3sig(unit / 0.3025) if unit else None
        _set(ws, r, 1, label, font=LABEL_FONT, border=True)
        _set(ws, r, 2, f"{total_r:,}円" if total_r else "", font=VALUE_FONT, border=True)
        _set(ws, r, 3, f"{unit_sqm_r:,}円" if unit_sqm_r else "", font=VALUE_FONT, border=True)
        _set(ws, r, 4, f"{unit_tsubo_r:,}円" if unit_tsubo_r else "", font=VALUE_FONT, border=True)
        r += 1
    r += 1

    # 比準表（建付減価列を削除した8列構成）
    if hijun_rows:
        _section_header(ws, r, "■ 比準表（標準画地の比準価格）")
        r += 1
        # 列構成（8列、建付減価削除済み）：
        # 1=事例番号, 2=取引価格, 3=事情補正, 4=時点修正,
        # 5=標準化補正, 6=地域格差, 7=試算値, 8=比準値
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        for j, h in enumerate(["事例番号", "取引価格(円/㎡)", "事情補正", "時点修正",
                               "標準化補正", "地域格差",
                               "試算値(円/㎡)", "比準値(円/㎡)"]):
            _set(ws, r, j+1, h, font=LABEL_FONT, fill=header_fill, border=True,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.row_dimensions[r].height = 32
        r += 1
        # 試算値の中央値を比準値に
        n_rows = len(hijun_rows)
        shisan_list = sorted(h["試算値"] for h in hijun_rows)
        if n_rows % 2 == 1:
            hijun_central = shisan_list[n_rows // 2]
        else:
            hijun_central = (shisan_list[n_rows // 2 - 1] + shisan_list[n_rows // 2]) / 2

        block_start_row = r
        center_align = Alignment(horizontal="center", vertical="center")
        # 表示順を [top2, top1, top3] に並び替え（規範性の高い事例を中央に配置、視覚強調なし）
        # hijun_rows は [top1, top2, top3] の順で来る
        if len(hijun_rows) == 3:
            display_rows = [hijun_rows[1], hijun_rows[0], hijun_rows[2]]
        elif len(hijun_rows) == 2:
            display_rows = [hijun_rows[1], hijun_rows[0]]
        else:
            display_rows = hijun_rows
        for idx, h in enumerate(display_rows):
            # 色強調なし（位置で識別：中央＝規範性の高い事例）
            fill = None
            font_top = VALUE_FONT
            label_font = LABEL_FONT
            top_row = r
            bot_row = r + 1
            # 補正項目の分子/分母（鑑定書様式）
            # 時点修正：分子側（査定時点 / 事例時点）
            # 標準化補正・地域格差：分母側（100 / 事例評点 = 事例側を分母に置く慣習）
            jijo_top, jijo_bot = _hijun_top_bottom(h["事情補正"], h.get("事情補正_適用", False))
            time_top, time_bot = _hijun_top_bottom(h["時点修正"], mode="top")
            hyo_top, hyo_bot = _hijun_top_bottom(h["標準化補正"], mode="bottom")
            chi_top, chi_bot = _hijun_top_bottom(h["地域格差"], mode="bottom")
            # 上行（分子）：列 3=事情補正, 4=時点修正, 5=標準化補正, 6=地域格差
            _set(ws, top_row, 3, jijo_top, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, top_row, 4, time_top, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, top_row, 5, hyo_top, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, top_row, 6, chi_top, font=font_top, fill=fill, border=True, align=center_align)
            # 下行（分母）
            _set(ws, bot_row, 3, jijo_bot, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, bot_row, 4, time_bot, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, bot_row, 5, hyo_bot, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, bot_row, 6, chi_bot, font=font_top, fill=fill, border=True, align=center_align)
            # 2行マージ：事例番号(1), 取引価格(2), 試算値(7)
            for col in [1, 2, 7]:
                ws.merge_cells(start_row=top_row, start_column=col,
                               end_row=bot_row, end_column=col)
            # 事例番号 = MLITデータ番号（透明性のため、人為的ラベルではない）
            case_no_str = str(h.get("事例番号", "?"))
            _set(ws, top_row, 1, case_no_str, font=label_font, fill=fill, border=True,
                 align=center_align)
            _set(ws, top_row, 2, f"{int(h['取引価格']):,}",
                 font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, top_row, 7, f"{int(round(h['試算値'])):,}",
                 font=font_top, fill=fill, border=True, align=center_align)
            r += 2
        block_end_row = r - 1
        # 比準値列（8列目）を全事例マージ
        ws.merge_cells(start_row=block_start_row, start_column=8,
                       end_row=block_end_row, end_column=8)
        _set(ws, block_start_row, 8, f"{int(round(hijun_central)):,}",
             font=Font(name="游ゴシック", size=12, bold=True, color="C00000"),
             border=True,
             align=Alignment(horizontal="center", vertical="center"))
        # 注釈
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        _set(ws, r, 1,
             "※ 事例番号 = MLITデータ原本の行番号。比準値 = 3事例の試算値の中央値。"
             "各補正は「分子/分母」形式（上段=分子、下段=分母）。「100/-」は補正非該当。"
             "標準化補正＝画地条件（規模, 形状, 方位, 袋地, 不整形）、"
             "地域格差＝地域・街路・交通要因（道路幅員, 駅徒歩, 容積率, 私道, 地区平均, 駅平均）のヘドニック係数積。"
             "**中央行（2段目）＝規範性の高い事例**（top1）、上下行は検証用の類似事例。",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 55
        r += 2

    # 比準表の内訳：取引事例の補修正率と地域格差率（鑑定実務標準フォーマット、9列）
    if hijun_detail_rows:
        _section_header(ws, r, "■ 比準表の内訳（取引事例の補修正率と地域格差率）",
                        end_col=9)
        r += 1

        # 2段ヘッダ：上段は「地域格差」のグループラベル
        hdr_fill = PatternFill("solid", fgColor="D9E1F2")
        col_labels_top = ["事例番号", "事情補正", "時点修正",
                          "標準化補正", "地域格差", "", "", "", ""]
        for j, h in enumerate(col_labels_top):
            _set(ws, r, j+1, h, font=LABEL_FONT, fill=hdr_fill, border=True,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        # 地域格差は4区分＋相乗積マージ（列5-9）— 相乗積=地域格差の積であることを明示
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
        # 単独列は2段マージ（縦）— 相乗積は地域格差サブヘッダ配下なので除外
        for col in [1, 2, 3, 4]:
            ws.merge_cells(start_row=r, start_column=col, end_row=r+1, end_column=col)
        r += 1
        # 下段：地域格差の4細目＋相乗積
        col_labels_bot = ["", "", "", "", "街路条件\n（総和）",
                          "交通接近条件\n（総和）", "環境条件\n（総和）", "行政的条件\n（総和）",
                          "相乗積\n（地域格差積）"]
        for j, h in enumerate(col_labels_bot):
            if h:  # 既にマージされていないセルのみ
                _set(ws, r, j+1, h, font=LABEL_FONT, fill=hdr_fill, border=True,
                     align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.row_dimensions[r-1].height = 18
        ws.row_dimensions[r].height = 30
        r += 1

        def _fmt_pct(v):
            """+X.X / ±0 / ▲X.X 形式"""
            if v is None:
                return "―"
            v = round(v, 1)
            if abs(v) < 0.05:
                return "±0"
            if v > 0:
                return f"+{v}"
            return f"▲{abs(v)}"

        def _filter_nonzero(items):
            """サブ項目のうち ±0（絶対値<0.05）を除外。
            v1.2.1: 「地区」エントリは β=0.81 と高インパクトの最重要要因なので、
            ±0でも常時表示する（事例と本物件の地区が同じことを白箱性として明示）。
            """
            keep = []
            for lbl, pct in items:
                if lbl.startswith("地区"):
                    keep.append((lbl, pct))  # 地区は常時表示
                elif abs(round(pct, 1)) >= 0.05:
                    keep.append((lbl, pct))
            return keep

        def _join_subitems(items, hide_zero=True):
            """[(label, pct), ...] を multi-line text に。±0は非表示（hide_zero=True）。"""
            if hide_zero:
                items = _filter_nonzero(items)
            if not items:
                return "標準的 ±0"
            return "\n".join(f"{lbl} {_fmt_pct(pct)}" for lbl, pct in items)

        cell_font = Font(name="游ゴシック", size=9)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 取引事例 行（事例番号で表示）— 中央に規範性の高い事例(top1)を配置
        if len(hijun_detail_rows) == 3:
            display_detail = [hijun_detail_rows[1], hijun_detail_rows[0], hijun_detail_rows[2]]
        elif len(hijun_detail_rows) == 2:
            display_detail = [hijun_detail_rows[1], hijun_detail_rows[0]]
        else:
            display_detail = hijun_detail_rows
        for idx, d in enumerate(display_detail):
            fill = None
            f = cell_font
            _set(ws, r, 1, d.get("事例番号", "?"), font=f, fill=fill, border=True, align=center_align)
            jijo_lbl, jijo_pct = d.get("事情補正", ("正常", 0.0))
            _set(ws, r, 2, f"{jijo_lbl}\n{_fmt_pct(jijo_pct)}", font=f, fill=fill, border=True, align=center_align)
            _set(ws, r, 3, _fmt_pct(d.get("時点修正_pct", 0)), font=f, fill=fill, border=True, align=center_align)
            std_items = _filter_nonzero(d.get("規模", []) + d.get("画地", []))
            std_text = "\n".join(f"{lbl} {_fmt_pct(pct)}" for lbl, pct in std_items) if std_items else "標準的 ±0"
            std_text += f"\n総和 {_fmt_pct(d.get('標準化補正_総和', 0))}"
            _set(ws, r, 4, std_text, font=f, fill=fill, border=True, align=center_align)
            street_text = _join_subitems(d.get("街路", []))
            street_text += f"\n総和 {_fmt_pct(d.get('街路_総和', 0))}"
            _set(ws, r, 5, street_text, font=f, fill=fill, border=True, align=center_align)
            tr_text = _join_subitems(d.get("交通接近", []))
            tr_text += f"\n総和 {_fmt_pct(d.get('交通接近_総和', 0))}"
            _set(ws, r, 6, tr_text, font=f, fill=fill, border=True, align=center_align)
            env_text = _join_subitems(d.get("環境", []))
            env_text += f"\n総和 {_fmt_pct(d.get('環境_総和', 0))}"
            _set(ws, r, 7, env_text, font=f, fill=fill, border=True, align=center_align)
            adm_text = _join_subitems(d.get("行政", []))
            adm_text += f"\n総和 {_fmt_pct(d.get('行政_総和', 0))}"
            _set(ws, r, 8, adm_text, font=f, fill=fill, border=True, align=center_align)
            _set(ws, r, 9, d.get("相乗積", 100), font=f, fill=fill, border=True, align=center_align)
            ws.row_dimensions[r].height = 70
            r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        koji_rate = rate_info.get("rate", 0) or 0
        _set(ws, r, 1,
             f"※ 時点修正率査定根拠：地価公示の年次変動率を参考に、地域の地価動向を分析の上、"
             f"年率 {koji_rate*100:+.1f}% で査定（{rate_info.get('method','')}, n = {rate_info.get('n_points', 0)} 地点）。"
             "各補正率は %-point 表記、相乗積は 100 を基準とする指数。"
             "**中央行（2段目）＝規範性の高い事例**（top1）、上下行は検証用の類似事例。"
             "標準化補正の細目（規模・形状・方位）と地域格差の細目（街路・交通接近・環境・行政）は、"
             "対応するヘドニック係数を反映。±0 の細目は非表示。",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 50
        r += 2

    # ■ 個別格差 + 査定価格の算定（業者用：縦並び個別格差 + 横並び算定式）
    if hijun_rows:
        # 規範性の高い事例（top1）の個別格差を採用
        primary_h_gy = next((h for h in hijun_rows if h.get("順位") == "規範性の高い事例"),
                            hijun_rows[0])
        center_align_gy = Alignment(horizontal="center", vertical="center")
        right_align_gy = Alignment(horizontal="right", vertical="center")

        # 標準画地の試算値（= 比準値 = 3事例の試算値の中央値）— 既に比準表で計算済み
        hijun_central_val = int(round(hijun_central))

        # 個別格差ブロック開始行
        kobetsu_start_row = r
        _set(ws, r, 1, "■ 個別格差",
             font=Font(name="游ゴシック", size=11, bold=True, color="FFFFFF"),
             fill=SECTION_FILL,
             align=Alignment(horizontal="left", vertical="center"))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        # 査定価格の算定ヘッダ（右側）
        _set(ws, r, 4, "■ 査定価格の算定",
             font=Font(name="游ゴシック", size=11, bold=True, color="FFFFFF"),
             fill=SECTION_FILL,
             align=Alignment(horizontal="left", vertical="center"))
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        r += 1

        def _fmt_kobetsu_v(v):
            iv = round(v, 1)
            return int(iv) if iv == int(iv) else iv

        # 個別格差 縦表（列A=ラベル、列B=値）
        # v1.2.1: target の属性をラベル括弧に転記（Style B）、target が中間画地の場合は角地行を非表示

        # 角地（target に角地補正率 > 0 が明示入力された場合のみ表示）
        target_kado_val_gy = primary_h_gy.get("個別格差_角地", 0)
        kado_row_gy = None  # 非表示の場合は None
        if abs(round(target_kado_val_gy, 1)) >= 0.05:
            _set(ws, r, 1, "角地（角地）", font=LABEL_FONT, border=True, align=center_align_gy)
            _set(ws, r, 2, _fmt_kobetsu_v(target_kado_val_gy),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            kado_row_gy = r
            r += 1

        # 算定式の視覚アンカー行（角地が非表示なら方位の行に揃える）
        first_kobetsu_row = kado_row_gy if kado_row_gy is not None else r

        # 方位（target の方位を表示ラベルに）
        target_dir_gy = str(target.get("前面道路:方位", "")).strip()
        houi_label_gy = f"方位（{target_dir_gy}）" if target_dir_gy else "方位"
        houi_val = _fmt_kobetsu_v(primary_h_gy.get("個別格差_方位", 0))
        _set(ws, r, 1, houi_label_gy, font=LABEL_FONT, border=True, align=center_align_gy)
        _set(ws, r, 2, houi_val, font=VALUE_FONT, border=True, align=center_align_gy)
        houi_row_gy = r
        r += 1

        # 不整形（target の土地形状を表示ラベルに、v1.2.1）
        target_shape_gy = str(target.get("土地の形状", "")).strip()
        fusei_label_gy = f"不整形（{target_shape_gy}）" if target_shape_gy else "不整形"
        fusei_val = _fmt_kobetsu_v(primary_h_gy.get("個別格差_不整形", 0))
        _set(ws, r, 1, fusei_label_gy, font=LABEL_FONT, border=True, align=center_align_gy)
        _set(ws, r, 2, fusei_val, font=VALUE_FONT, border=True, align=center_align_gy)
        fusei_row_gy = r
        r += 1

        # 総和（Excel関数式）— 表示中の格差行のみを積算
        _set(ws, r, 1, "総和", font=LABEL_FONT, fill=PatternFill("solid", fgColor="FFF2CC"),
             border=True, align=center_align_gy)
        factor_refs_gy = []
        if kado_row_gy is not None:
            factor_refs_gy.append(f"(100+B{kado_row_gy})/100")
        factor_refs_gy.append(f"(100+B{houi_row_gy})/100")
        factor_refs_gy.append(f"(100+B{fusei_row_gy})/100")
        soan_formula_gy = "=" + "*".join(factor_refs_gy) + "*100"
        soan_cell_gy = ws.cell(row=r, column=2, value=soan_formula_gy)
        soan_cell_gy.font = Font(name="游ゴシック", size=10, bold=True)
        soan_cell_gy.border = BORDER
        soan_cell_gy.alignment = center_align_gy
        soan_cell_gy.number_format = "0.00"
        soan_cell_gy.fill = PatternFill("solid", fgColor="FFF2CC")
        soan_row_gy = r
        r += 1

        # ====== 査定価格の算定 行（個別格差ブロックの隣、first_kobetsu_row 行から横並び）======
        # 算定式の表示行：D=試算値, E="×", F=総和参照, G="÷ 100 ≒", H=案件査定価格
        # v1.2.1: 視覚アンカーは first_kobetsu_row（角地非表示時は方位の行）

        # 試算値 (D列、first_kobetsu_row 行)
        _set(ws, first_kobetsu_row, 4, hijun_central_val,
             font=Font(name="游ゴシック", size=12, bold=True),
             border=True, align=center_align_gy, number_format="#,##0")
        _set(ws, first_kobetsu_row+1, 4, "標準画地の試算値(円/㎡)",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=center_align_gy)

        # × 演算子 (E列)
        _set(ws, first_kobetsu_row, 5, "×",
             font=Font(name="游ゴシック", size=14, bold=True),
             align=center_align_gy)

        # 総和/100 (F列) — 分子=B{soan_row_gy}, 分母=100 を 2行縦に表示
        soan_ref_cell = ws.cell(row=first_kobetsu_row, column=6, value=f"=B{soan_row_gy}")
        soan_ref_cell.font = Font(name="游ゴシック", size=11, bold=True)
        soan_ref_cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=Side(border_style="thin", color="000000"))
        soan_ref_cell.alignment = center_align_gy
        soan_ref_cell.number_format = "0.00"
        denom_cell = ws.cell(row=first_kobetsu_row+1, column=6, value=100)
        denom_cell.font = Font(name="游ゴシック", size=11, bold=True)
        denom_cell.border = Border(left=THIN, right=THIN, top=Side(border_style="thin", color="000000"), bottom=THIN)
        denom_cell.alignment = center_align_gy

        # ≒ (G列)
        _set(ws, first_kobetsu_row, 7, "≒",
             font=Font(name="游ゴシック", size=14, bold=True),
             align=center_align_gy)

        # 案件査定価格 (H列) — Excel関数式
        anken_inner = f"D{first_kobetsu_row}*B{soan_row_gy}"
        anken_formula_gy = f"=ROUND({anken_inner},-(LEN(INT({anken_inner}))-3))/100"
        anken_cell_gy = ws.cell(row=first_kobetsu_row, column=8, value=anken_formula_gy)
        anken_cell_gy.font = Font(name="游ゴシック", size=14, bold=True, color="C00000")
        anken_cell_gy.border = BORDER
        anken_cell_gy.alignment = center_align_gy
        anken_cell_gy.number_format = "#,##0"
        _set(ws, first_kobetsu_row+1, 8, "案件査定価格(円/㎡)",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=center_align_gy)

        # 個別格差 + 査定価格の算定 ブロックの後は r が総和の次に進んでいる
        r += 1  # 空行

        # 注釈
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        _set(ws, r, 1,
             "※ 個別格差は規範性の高い事例（top1）と本物件の差から算出。"
             "**角地補正は業者の入力値（デフォルト 0%）**。"
             "MLITデータに角地情報が無いためヘドニックで推定不能 → 白箱ポリシー上、自動値は与えず業者判断に委ねる。"
             "方位・不整形補正はヘドニック係数 β（dir_score, D_fuseikei）に基づく "
             "exp(β×(本物件 − 事例)) − 1。"
             "総和 = (1 + 角地/100) × (1 + 方位/100) × (1 + 不整形/100) × 100。"
             "案件査定価格 = 標準画地の試算値 × 総和 / 100（上位3桁四捨五入）。",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 45
        r += 2

        # ■ 取引事例の概要（横並び、3事例の詳細データ）
        _section_header(ws, r, "■ 取引事例の概要", end_col=12)
        r += 1
        gaiyo_headers = ["事例番号", "取引㎡単価", "取引時点", "地区", "最寄り駅",
                         "駅距離(分)", "道路", "道路幅員(m)", "方位", "形状",
                         "地積(㎡)", "用途地域", "容積率(%)"]
        hdr_fill_g = PatternFill("solid", fgColor="D9E1F2")
        # 13列に拡張、セクションヘッダのマージも更新
        ws.unmerge_cells(start_row=r-1, start_column=1, end_row=r-1, end_column=12)
        ws.merge_cells(start_row=r-1, start_column=1, end_row=r-1, end_column=13)
        for j, h in enumerate(gaiyo_headers):
            _set(ws, r, j+1, h, font=LABEL_FONT, fill=hdr_fill_g, border=True,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.row_dimensions[r].height = 30
        r += 1

        # 表示順は比準表と整合：[top2, top1, top3]（中央=規範性の高い事例）
        if len(hijun_rows) == 3:
            gaiyo_display = [hijun_rows[1], hijun_rows[0], hijun_rows[2]]
        elif len(hijun_rows) == 2:
            gaiyo_display = [hijun_rows[1], hijun_rows[0]]
        else:
            gaiyo_display = hijun_rows

        def _fmt_or_dash(v, kind="num"):
            """欠損値は ― で表示。"""
            if v is None or v == "" or (isinstance(v, float) and pd.isna(v)):
                return "―"
            if kind == "int":
                try:
                    return int(v)
                except (TypeError, ValueError):
                    return str(v)
            if kind == "num":
                try:
                    return f"{int(round(float(v))):,}"
                except (TypeError, ValueError):
                    return str(v)
            return str(v)

        for h in gaiyo_display:
            _set(ws, r, 1, str(h.get("事例番号", "?")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 2, _fmt_or_dash(h.get("取引価格"), "num"),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 3, str(h.get("取引四半期", "") or h.get("取引時点", "")),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 4, str(h.get("地区", "")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 5, str(h.get("最寄駅", "")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 6, _fmt_or_dash(h.get("駅距離"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 7, str(h.get("道路種別", "")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 8, _fmt_or_dash(h.get("道路幅員"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 9, str(h.get("方位", "")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 10, str(h.get("形状", "")) or "—", font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 11, _fmt_or_dash(h.get("面積"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 12, str(h.get("用途地域", "")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 13, _fmt_or_dash(h.get("容積率_pct"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            r += 1

        # 注釈
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        _set(ws, r, 1,
             "※ 事例番号 = MLITデータ原本の行番号。**中央行＝規範性の高い事例**（top1）。"
             "取引時点は四半期表記（例：2025年第2四半期）。",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 30
        r += 2

    # ■ 公示価格の概要（業者用：取引事例の概要と同様の横並びテーブル）
    # 地域標準価格チェックで選定された公示標準地の詳細属性を表示
    koji_points_for_summary = standard_check.get("selected_points", []) if standard_check else []
    if koji_points_for_summary:
        _section_header(ws, r, "■ 公示価格の概要", end_col=12)
        r += 1
        koji_headers = ["公示番号", "公示価格(円/㎡)", "所在", "地区", "最寄駅",
                        "駅距離(m)", "道路", "道路幅員(m)", "方位", "形状",
                        "地積(㎡)", "用途地域", "容積率(%)"]
        # 13列に拡張、セクションヘッダのマージも更新
        ws.unmerge_cells(start_row=r-1, start_column=1, end_row=r-1, end_column=12)
        ws.merge_cells(start_row=r-1, start_column=1, end_row=r-1, end_column=13)
        hdr_fill_kj = PatternFill("solid", fgColor="D9E1F2")
        for j, h in enumerate(koji_headers):
            _set(ws, r, j+1, h, font=LABEL_FONT, fill=hdr_fill_kj, border=True,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.row_dimensions[r].height = 30
        r += 1

        def _fmt_or_dash_k(v, kind="num"):
            if v is None or v == "" or v == "_":
                return "―"
            try:
                if isinstance(v, float) and pd.isna(v):
                    return "―"
            except (TypeError, ValueError):
                pass
            if kind == "int":
                try:
                    return int(float(v))
                except (TypeError, ValueError):
                    return str(v)
            if kind == "num":
                try:
                    return f"{int(round(float(v))):,}"
                except (TypeError, ValueError):
                    return str(v)
            return str(v)

        center_align_kj = Alignment(horizontal="center", vertical="center")
        # 通常は1地点（場所による価格水準差を排除するため類似度スコアで絞込み）
        for pt in koji_points_for_summary[:5]:
            _set(ws, r, 1, _short_koji_id(str(pt.get("id", ""))),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 2, _fmt_or_dash_k(pt.get("price_at_asof"), "num"),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 3, _short_koji_addr(str(pt.get("address", "")),
                                             str(pt.get("district", ""))),
                 font=VALUE_FONT, border=True,
                 align=Alignment(horizontal="center", vertical="center"))
            _set(ws, r, 4, str(pt.get("district", "")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 5, str(pt.get("station", "")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 6, _fmt_or_dash_k(pt.get("station_dist_m"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 7, str(pt.get("road_type", "")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 8, _fmt_or_dash_k(pt.get("road_width"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 9, str(pt.get("road_dir", "")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 10, _koji_shape_label(pt.get("frontage_ratio"),
                                              pt.get("depth_ratio")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 11, _fmt_or_dash_k(pt.get("area_sqm"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 12, str(pt.get("zoning", "")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 13, _fmt_or_dash_k(pt.get("floor_area_ratio"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            r += 1

        # 注釈
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        _set(ws, r, 1,
             "※ 公示番号 = 「市区町村-連番」形式（例：世田谷-50 = 13112-000-050）。"
             "公示価格は査定時点へ線形補間済み。"
             "形状は間口比率(L01_036)・奥行比率(L01_037)から推定（最大比 ≤1.5: 整形、≤2.5: やや細長、それ以上: 細長）。",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 40
        r += 2

    # 公示地価の時系列推移（折れ線グラフ）— 時点修正に使用した標準地と整合
    koji_ts_obj = ctx.get("koji_timeseries", {})
    if isinstance(koji_ts_obj, dict):
        koji_ts = koji_ts_obj.get("data", [])
        koji_label = koji_ts_obj.get("label", "")
    else:
        koji_ts = koji_ts_obj
        koji_label = ""
    if len(koji_ts) >= 2:
        ts_header_row = r
        _set(ws, r, 1,
             f"▼ {koji_label} の直近5年間の価格推移",
             font=Font(name="游ゴシック", size=10, bold=True, color="595959"))
        r += 1
        for j, h in enumerate(["評価年", "平均単価 (円/㎡)"]):
            _set(ws, r, j+1, h, font=LABEL_FONT,
                 fill=PatternFill("solid", fgColor="D9E1F2"), border=True,
                 align=Alignment(horizontal="center", vertical="center"))
        r += 1
        ts_data_start = r
        for pt in koji_ts:
            _set(ws, r, 1, pt["year"], font=VALUE_FONT, border=True,
                 align=Alignment(horizontal="center", vertical="center"))
            _set(ws, r, 2, pt["price"], font=VALUE_FONT, border=True,
                 number_format='#,##0',
                 align=Alignment(horizontal="right", vertical="center"))
            r += 1
        ts_data_end = r - 1

        # 折れ線グラフ（内部タイトルなし。グラフシートの section header に統一）
        line = LineChart()
        line.title = None
        line.legend = None
        line.height = 7
        line.width = 14
        data_ref = Reference(ws, min_col=2, min_row=ts_data_start, max_col=2, max_row=ts_data_end)
        cats_ref = Reference(ws, min_col=1, min_row=ts_data_start, max_col=1, max_row=ts_data_end)
        line.add_data(data_ref, titles_from_data=False)
        line.set_categories(cats_ref)
        line.y_axis.title = "単価 (円/㎡)"
        line.x_axis.title = "評価年"
        # Y軸範囲と目盛を明示設定（折れ線が中央に来るよう、かつ目盛を表示）
        prices = [pt["price"] for pt in koji_ts]
        y_min = min(prices)
        y_max = max(prices)
        if y_max > y_min:
            margin = (y_max - y_min) * 0.3
        else:
            margin = y_max * 0.05
        axis_min = max(0, y_min - margin)
        axis_max = y_max + margin
        line.y_axis.scaling.min = axis_min
        line.y_axis.scaling.max = axis_max
        # 目盛間隔を 5分割に
        line.y_axis.majorUnit = (axis_max - axis_min) / 5
        line.y_axis.delete = False
        line.y_axis.majorTickMark = 'out'
        line.y_axis.number_format = '#,##0'
        # データラベル：年 + 値 を併記
        dl_line = DataLabelList()
        dl_line.showVal = True
        dl_line.showCatName = True
        dl_line.showSerName = False
        dl_line.showLegendKey = False
        dl_line.position = 't'
        dl_line.separator = '\n'
        line.dataLabels = dl_line
        line.x_axis.delete = False
        line.series[0].smooth = False
        # グラフはグラフ専用シートに配置
        graph_ws_ref = ctx.get('_graph_ws')
        if graph_ws_ref is not None:
            gr = ctx.get('_graph_row', 3)
            # セクション見出し
            _set(graph_ws_ref, gr, 1,
                 f"■ {koji_label} の直近5年間の価格推移",
                 font=Font(name="游ゴシック", size=11, bold=True, color="FFFFFF"),
                 fill=SECTION_FILL,
                 align=Alignment(horizontal="left", vertical="center"))
            graph_ws_ref.merge_cells(start_row=gr, start_column=1, end_row=gr, end_column=14)
            graph_ws_ref.add_chart(line, f"A{gr+1}")
            ctx['_graph_row'] = gr + 17  # line chart 7cm ≈ 15行 + buffer
        else:
            ws.add_chart(line, f"D{ts_header_row}")
    r += 1

    # ヘドニック回帰サマリ + β符号チェック（末尾：技術詳細・係数全開示の参考情報）
    _section_header(ws, r, "■ ヘドニック回帰サマリ（係数全開示・参考情報）")
    r += 1
    if hed["ok"]:
        _set(ws, r, 1, f"サンプル数 n = {hed['n']}", font=VALUE_FONT)
        _set(ws, r, 3, f"R² = {hed['r2']:.3f}", font=VALUE_FONT)
        _set(ws, r, 5, f"自由度調整済 R² = {hed['adj_r2']:.3f}", font=VALUE_FONT)
        r += 1
        for j, h in enumerate(["特徴量", "推定値 β", "標準誤差", "p値", "有意性"]):
            _set(ws, r, j+1, h, font=LABEL_FONT, fill=PatternFill("solid", fgColor="D9E1F2"), border=True)
        r += 1
        coef_data_start = r  # 表での開始行（チャートのカテゴリ範囲開始）
        for name, c in hed["coefficients"].items():
            if name == "const":
                continue  # グラフから定数項は除外
            p = c["p"]
            if p < 0.05: fill = P_LOW_FILL; sig = "** (p<0.05)"
            elif p < 0.10: fill = P_MID_FILL; sig = "*  (p<0.10)"
            else: fill = P_HIGH_FILL; sig = "ns"
            # 表示用：業者用シート column 1〜5（簡潔ラベル + 数値）
            _set(ws, r, 1, c["label"], font=VALUE_FONT, border=True)
            _set(ws, r, 2, float(c['beta']), font=VALUE_FONT, border=True,
                 number_format='+0.0000;-0.0000;0.0000')
            _set(ws, r, 3, float(c['se']), font=VALUE_FONT, border=True,
                 number_format='0.0000')
            _set(ws, r, 4, float(p), font=VALUE_FONT, border=True, fill=fill,
                 number_format='0.0000')
            _set(ws, r, 5, sig, font=VALUE_FONT, border=True, fill=fill)
            r += 1
        coef_data_end = r - 1
        # 定数項を末尾に追加（参考表示、グラフ対象外）
        if "const" in hed["coefficients"]:
            c = hed["coefficients"]["const"]
            p = c["p"]
            if p < 0.05: fill = P_LOW_FILL; sig = "** (p<0.05)"
            elif p < 0.10: fill = P_MID_FILL; sig = "*  (p<0.10)"
            else: fill = P_HIGH_FILL; sig = "ns"
            _set(ws, r, 1, c["label"], font=VALUE_FONT, border=True)
            _set(ws, r, 2, float(c['beta']), font=VALUE_FONT, border=True,
                 number_format='+0.0000;-0.0000;0.0000')
            _set(ws, r, 3, float(c['se']), font=VALUE_FONT, border=True,
                 number_format='0.0000')
            _set(ws, r, 4, float(p), font=VALUE_FONT, border=True, fill=fill,
                 number_format='0.0000')
            _set(ws, r, 5, sig, font=VALUE_FONT, border=True, fill=fill)
            r += 1
        r += 1

        # ヘドニック係数 棒グラフ（白箱AVM の象徴：全特徴量の β を可視化）
        bar = BarChart()
        bar.type = "bar"  # 横向き棒グラフ（特徴量名が長いので）
        bar.style = 11
        bar.title = None  # グラフシートの section header に統一
        bar.legend = None
        bar.height = 10  # cm
        bar.width = 16   # cm
        # チャートは業者用シートの column 1（特徴量名）と column 2（β値）を直接参照
        data_ref = Reference(ws, min_col=2, min_row=coef_data_start,
                             max_col=2, max_row=coef_data_end)
        cats_ref = Reference(ws, min_col=1, min_row=coef_data_start,
                             max_col=1, max_row=coef_data_end)
        bar.add_data(data_ref, titles_from_data=False)
        bar.set_categories(cats_ref)
        bar.y_axis.title = None
        bar.x_axis.title = "係数 β（負＝単価↓、正＝単価↑）"
        bar.y_axis.delete = False
        # データラベル：カテゴリ名 + 値 を表示（Y軸ラベルが Excel で表示されない問題への対処）
        dl = DataLabelList()
        dl.showVal = True
        dl.showCatName = True   # 棒の右側に「面積 +0.0729」形式で表示
        dl.showSerName = False
        dl.showLegendKey = False
        dl.position = 'outEnd'
        dl.separator = ' '
        bar.dataLabels = dl
        # グラフはグラフ専用シートに配置
        graph_ws_ref = ctx.get('_graph_ws')
        if graph_ws_ref is not None:
            gr = ctx.get('_graph_row', 3)
            _set(graph_ws_ref, gr, 1,
                 "■ ヘドニック回帰係数 β（単価への影響度・対数空間）",
                 font=Font(name="游ゴシック", size=11, bold=True, color="FFFFFF"),
                 fill=SECTION_FILL,
                 align=Alignment(horizontal="left", vertical="center"))
            graph_ws_ref.merge_cells(start_row=gr, start_column=1, end_row=gr, end_column=14)
            graph_ws_ref.add_chart(bar, f"A{gr+1}")
            ctx['_graph_row'] = gr + 24  # bar chart 10cm ≈ 22行 + buffer
        else:
            anchor_cell = f"G{coef_data_start}"
            ws.add_chart(bar, anchor_cell)

        # β符号チェック（期待符号 vs 実際の符号）
        EXPECTED_SIGNS = {
            "ln_area": ("負", "面積大→単価下落"),
            "walk_min": ("負", "駅遠→単価下落"),
            "D_shidou": ("負", "私道→減価"),
            "D_fukuro": ("負", "袋地→減価"),
            "D_fuseikei": ("負", "不整形→減価"),
        }
        coef = hed["coefficients"]
        sign_check_label = Font(name="游ゴシック", size=10, bold=True, color="595959")
        _set(ws, r, 1, "▼ β符号チェック（期待符号 vs 実際）", font=sign_check_label)
        r += 1
        for j, h in enumerate(["特徴量", "期待符号", "実際 β", "整合", "経済的解釈"]):
            _set(ws, r, j+1, h, font=LABEL_FONT,
                 fill=PatternFill("solid", fgColor="D9E1F2"), border=True)
        r += 1
        any_significant_inconsistent = False
        for name, (expected, interpretation) in EXPECTED_SIGNS.items():
            if name not in coef:
                continue
            beta = coef[name]["beta"]
            p = coef[name]["p"]
            is_neg_expected = (expected == "負")
            is_consistent = (is_neg_expected and beta < 0) or (not is_neg_expected and beta > 0)
            is_significant = p < 0.10  # p<0.10 で統計的に有意
            if not is_consistent and is_significant:
                any_significant_inconsistent = True
                mark = "× 反転（有意・要確認）"
                ok_fill = P_HIGH_FILL
            elif not is_consistent:
                mark = "△ 反転（非有意・ノイズ範囲）"
                ok_fill = P_MID_FILL
            else:
                mark = "○ 整合"
                ok_fill = P_LOW_FILL
            _set(ws, r, 1, coef[name]["label"], font=VALUE_FONT, border=True)
            _set(ws, r, 2, expected, font=VALUE_FONT, border=True)
            _set(ws, r, 3, f"{beta:+.4f}", font=VALUE_FONT, border=True)
            _set(ws, r, 4, mark, font=VALUE_FONT, border=True, fill=ok_fill)
            _set(ws, r, 5, interpretation, font=VALUE_FONT, border=True)
            r += 1
        if any_significant_inconsistent:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            _set(ws, r, 1,
                 "※ 統計的に有意な符号反転（p<0.10）は外れ値・特徴量不足・地区特性などの構造問題の可能性。事例を再確認してください。"
                 "非有意な反転（△）はノイズ範囲内のため実害なし。",
                 font=VALUE_FONT, fill=WARN_FILL)
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        _set(ws, r, 1, f"※ {hed['skip_reason']}（類似度ベース集約に降格）",
             font=VALUE_FONT, fill=WARN_FILL)
        r += 1
    r += 1

    # 散布図：全事例の駅距離 vs 時点修正後単価（査定価格と top3 を強調）
    adjusted_full = ctx.get("adjusted_full")
    if adjusted_full is not None and len(adjusted_full) >= 10:
        scatter_header_row = r
        _set(ws, r, 1, "▼ 散布図：駅距離 vs 単価（全事例・比較事例・査定価格）",
             font=Font(name="游ゴシック", size=10, bold=True, color="595959"))
        r += 1

        # データを 列 18-23 に書き込み（右側に隠れる、列幅も狭く）
        # 18=全事例 X, 19=全事例 Y, 20=top3 X, 21=top3 Y, 22=対象 X, 23=対象 Y
        scoped_data = []
        for _, rw in adjusted_full.iterrows():
            walk = rw.get("walk_min")
            price = rw.get("adjusted_unit_price") if "adjusted_unit_price" in rw else rw.get("unit_price")
            try:
                if walk is not None and not pd.isna(walk) and price is not None and not pd.isna(price):
                    scoped_data.append((float(walk), float(price)))
            except (TypeError, ValueError):
                pass

        # ヘッダ行（列幅縮小用に色だけつけて値は書かない）
        scatter_data_start = r
        for i, (walk, price) in enumerate(scoped_data):
            ws.cell(row=scatter_data_start + i, column=18, value=walk)
            ws.cell(row=scatter_data_start + i, column=19, value=price)
        scatter_scoped_end = scatter_data_start + len(scoped_data) - 1

        # top3 を column 20-21 に
        top3_data = []
        for _, rw in cases.iterrows():
            walk = rw.get("walk_min")
            price = rw.get("corrected_unit_price")
            if price is None or pd.isna(price):
                price = rw.get("adjusted_unit_price") or rw.get("unit_price")
            try:
                if walk is not None and not pd.isna(walk) and price is not None and not pd.isna(price):
                    top3_data.append((float(walk), float(price)))
            except (TypeError, ValueError):
                pass
        for i, (walk, price) in enumerate(top3_data):
            ws.cell(row=scatter_data_start + i, column=20, value=walk)
            ws.cell(row=scatter_data_start + i, column=21, value=price)
        scatter_top3_end = scatter_data_start + max(0, len(top3_data) - 1)

        # 査定価格を column 22-23 に（1点）
        target_walk = target.get("最寄駅:距離(分)")
        target_price = assess.get("central_unit_price")
        target_present = False
        if target_walk is not None and target_price is not None:
            try:
                ws.cell(row=scatter_data_start, column=22, value=float(target_walk))
                ws.cell(row=scatter_data_start, column=23, value=float(target_price))
                target_present = True
            except (TypeError, ValueError):
                pass

        # データ列の幅を狭く（右に隠す）
        for col_letter in ['R', 'S', 'T', 'U', 'V', 'W']:
            ws.column_dimensions[col_letter].width = 3

        # 散布図（マーカーのみ、線なし）
        from openpyxl.chart.marker import Marker
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        from openpyxl.drawing.fill import ColorChoice

        sc = ScatterChart()
        sc.title = None  # グラフシートの section header に統一
        sc.style = 13
        sc.height = 9
        sc.width = 16
        sc.scatterStyle = "marker"  # 線なし、マーカーのみ
        sc.x_axis.title = "最寄駅徒歩(分)"
        sc.y_axis.title = "時点修正後単価 (円/㎡)"

        def _styled_series(y_ref, x_ref, title, color, size, symbol='circle'):
            ser = Series(y_ref, x_ref, title=title)
            # 線を非表示
            ser.graphicalProperties = GraphicalProperties()
            ser.graphicalProperties.line = LineProperties(noFill=True)
            # マーカー設定
            mk = Marker(symbol=symbol, size=size)
            mk.graphicalProperties = GraphicalProperties(solidFill=color)
            mk.graphicalProperties.line = LineProperties(solidFill=color)
            ser.marker = mk
            return ser

        # Series 1: 全事例（青小マーカー）
        if scoped_data:
            x_all = Reference(ws, min_col=18, min_row=scatter_data_start, max_col=18, max_row=scatter_scoped_end)
            y_all = Reference(ws, min_col=19, min_row=scatter_data_start, max_col=19, max_row=scatter_scoped_end)
            sc.series.append(_styled_series(y_all, x_all, "全事例", "4472C4", 4, 'circle'))
        # Series 2: top3（赤大マーカー）
        if top3_data:
            x_t3 = Reference(ws, min_col=20, min_row=scatter_data_start, max_col=20, max_row=scatter_top3_end)
            y_t3 = Reference(ws, min_col=21, min_row=scatter_data_start, max_col=21, max_row=scatter_top3_end)
            sc.series.append(_styled_series(y_t3, x_t3, "比較事例top3", "C00000", 9, 'diamond'))
        # Series 3: 査定価格（緑★大マーカー）
        if target_present:
            x_tg = Reference(ws, min_col=22, min_row=scatter_data_start, max_col=22, max_row=scatter_data_start)
            y_tg = Reference(ws, min_col=23, min_row=scatter_data_start, max_col=23, max_row=scatter_data_start)
            sc.series.append(_styled_series(y_tg, x_tg, "査定価格", "00B050", 14, 'star'))

        # グラフはグラフ専用シートに配置
        graph_ws_ref = ctx.get('_graph_ws')
        if graph_ws_ref is not None:
            gr = ctx.get('_graph_row', 3)
            _set(graph_ws_ref, gr, 1,
                 "■ 散布図：駅距離 vs 単価（全事例青、比較事例top3赤、査定価格緑）",
                 font=Font(name="游ゴシック", size=11, bold=True, color="FFFFFF"),
                 fill=SECTION_FILL,
                 align=Alignment(horizontal="left", vertical="center"))
            graph_ws_ref.merge_cells(start_row=gr, start_column=1, end_row=gr, end_column=14)
            graph_ws_ref.add_chart(sc, f"A{gr+1}")
            ctx['_graph_row'] = gr + 21  # scatter chart 9cm ≈ 19行 + buffer
        else:
            ws.add_chart(sc, f"G{scatter_header_row}")
        # スキップして次のセクションへ進む（データ書き込みは右側列なので r は変えない）
        r += 2

    _adjust_col_widths(ws, [14, 10, 12, 16, 12, 14, 12, 16, 10, 10, 10, 10, 12, 10])


# ===== 顧客用シート =====
def _qualitative(beta_label, target_v, mean_v):
    """個別要因の定性表現（係数を開示せず方向と程度のみ）。"""
    if mean_v == 0 or mean_v is None or target_v is None:
        return None
    diff = target_v - mean_v
    if abs(diff) < 0.1 * abs(mean_v):
        return "同水準"
    if diff > 0:
        return "やや広め" if "面積" in beta_label else "やや遠め" if "駅" in beta_label else "ややプラス要因"
    return "やや狭め" if "面積" in beta_label else "やや近め" if "駅" in beta_label else "ややマイナス要因"


def _write_kokyaku_sheet(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("顧客用")
    target = ctx["target"]
    asof = ctx["asof"]
    rate_info = ctx["rate_info"]
    cases = ctx["cases"]
    assess = ctx["assess"]
    standard_check = ctx["standard_check"]
    is_degraded = not ctx["hedonic"]["ok"]
    hijun_rows = ctx.get("hijun_rows", [])

    # 印刷ヘッダ・フッタは _apply_page_setup() で統一設定（OEM 想定でブランド名は付けない）

    r = 1
    # タイトル（降格時はラベル変更）
    chome = target.get("丁目", "")
    location = f"{target['市区町村名']} {target.get('地区名', '')}{chome}"
    title_text = (
        f"土地価格 参考情報 — {location}"
        if is_degraded
        else f"土地机上査定書 — {location}"
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    assert_clean(title_text, "title")
    _set(ws, r, 1, title_text, font=TITLE_FONT, fill=TITLE_FILL,
         align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[r].height = 28
    r += 1

    # （OEM 想定のため自動生成バナーは出さない。提供者表示は印刷フッタや別途運用で）

    # 降格時の参考情報バナー（赤色）
    if is_degraded:
        warn_red_fill = PatternFill("solid", fgColor="C00000")
        warn_red_font = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        banner_text = (
            "※ 取引事例件数が不足しているため、本資料は「参考情報」としてご覧ください（机上査定書ではありません）。"
            " 正式な査定価格は、ご担当者による現地確認・追加調査を経て決定する必要があります。"
        )
        assert_clean(banner_text, "degraded banner")
        _set(ws, r, 1, banner_text, font=warn_red_font, fill=warn_red_fill,
             align=Alignment(wrap_text=True, vertical="center"))
        ws.row_dimensions[r].height = 38
        r += 2

    # ■ 査定結果サマリ
    _section_header(ws, r, "■ 査定結果サマリ", end_col=6)
    r += 1
    rng = assess["range"]
    target_area = target["面積(㎡)"]

    # 査定価格（総額＋㎡単価＋坪単価併記、降格時はラベル変更）
    price_label = "参考価格" if is_degraded else "査定価格"
    summary_rows = [
        (price_label, _format_price_full(rng["central_total"], target_area)),
        ("面積", f"{target_area} ㎡"),
        ("査定時点", asof.isoformat()),
    ]
    for label, value in summary_rows:
        assert_clean(label, "summary label")
        assert_clean(value, "summary value")
        _set(ws, r, 1, label, font=LABEL_FONT, border=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        _set(ws, r, 2, value, font=VALUE_FONT, border=True)
        r += 1
    r += 1

    # 価格レンジ（上限／中央／下限）
    _section_header(ws, r, "■ 価格レンジ", end_col=6)
    r += 1
    for j, h in enumerate(["区分", "総額", "㎡単価", "坪単価"]):
        _set(ws, r, j+1, h, font=LABEL_FONT,
             fill=PatternFill("solid", fgColor="D9E1F2"), border=True)
    r += 1
    for label, total, unit in [
        ("上限", rng["high_total"], rng["high_unit"]),
        ("中央", rng["central_total"], rng["central_unit"]),
        ("下限", rng["low_total"], rng["low_unit"]),
    ]:
        total_r = _round_3sig(total) if total else None
        unit_sqm_r = _round_3sig(unit) if unit else None
        unit_tsubo_r = _round_3sig(unit / 0.3025) if unit else None
        _set(ws, r, 1, label, font=LABEL_FONT, border=True)
        _set(ws, r, 2, f"{total_r:,}円" if total_r else "", font=VALUE_FONT, border=True)
        _set(ws, r, 3, f"{unit_sqm_r:,}円" if unit_sqm_r else "", font=VALUE_FONT, border=True)
        _set(ws, r, 4, f"{unit_tsubo_r:,}円" if unit_tsubo_r else "", font=VALUE_FONT, border=True)
        r += 1
    r += 1

    # ■ 標準価格
    _section_header(ws, r, "■ 標準価格", end_col=6)
    r += 1
    if standard_check.get("standard_price_per_sqm"):
        label = standard_check.get("label", "")
        if label:
            text = (f"本地区の{standard_check['source']}による標準価格"
                    f"（{label}）は {int(standard_check['standard_price_per_sqm']):,} 円/㎡ です。")
        else:
            text = f"本地区の{standard_check['source']}による標準価格は {int(standard_check['standard_price_per_sqm']):,} 円/㎡ です。"
    else:
        text = "本地区の公的な標準価格データが取得できなかったため、参考情報なしで査定しています。"
    assert_clean(text, "standard")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws, r, 1, text, font=VALUE_FONT, align=Alignment(wrap_text=True, vertical="top"))
    ws.row_dimensions[r].height = 30
    r += 2

    # ■ 時点修正
    _section_header(ws, r, "■ 時点修正", end_col=6)
    r += 1
    if rate_info.get("rate") is not None:
        rate_pct = rate_info["rate"] * 100
        direction = "上昇" if rate_pct > 0 else "下落" if rate_pct < 0 else "横ばい"
        text = (f"直近の本地区の地価は年率 {rate_pct:+.2f}% で{direction}しています。"
                f"この動きを踏まえて、過去の取引事例を査定時点に補正しています。")
    else:
        text = "時点修正の根拠となる地価データが取得できなかったため、補正なしで査定しています。"
    assert_clean(text, "time_adjust")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws, r, 1, text, font=VALUE_FONT, align=Alignment(wrap_text=True, vertical="top"))
    ws.row_dimensions[r].height = 36
    r += 2

    # ■ 公示価格の詳細（顧客用：縦並びで読みやすく）
    koji_points_k = standard_check.get("selected_points", []) if standard_check else []
    if koji_points_k:
        pt0 = koji_points_k[0]
        _section_header(ws, r, "■ 公示価格の詳細", end_col=6)
        r += 1
        short_id_k = _short_koji_id(str(pt0.get("id", "")))
        short_addr_k = _short_koji_addr(str(pt0.get("address", "")),
                                         str(pt0.get("district", "")))
        shape_k = _koji_shape_label(pt0.get("frontage_ratio"), pt0.get("depth_ratio"))
        price_k = pt0.get("price_at_asof")
        # 整形した項目リスト
        koji_info_lines = [
            ("公示番号", short_id_k),
            ("公示価格", f"{int(price_k):,} 円/㎡" if price_k else "—"),
            ("所在", short_addr_k or "—"),
            ("最寄駅", f"{pt0.get('station','')} （{int(pt0.get('station_dist_m') or 0)} m）"
                if pt0.get("station") else "—"),
            ("前面道路", f"{pt0.get('road_type','')} 幅員{int(pt0.get('road_width') or 0)}m "
                f"{pt0.get('road_dir','')}向"
                if pt0.get("road_type") else "—"),
            ("形状", shape_k),
            ("地積", f"{int(pt0.get('area_sqm') or 0)} ㎡" if pt0.get("area_sqm") else "—"),
            ("用途地域", pt0.get("zoning", "") or "—"),
            ("容積率", f"{int(pt0.get('floor_area_ratio') or 0)}%"
                if pt0.get("floor_area_ratio") else "—"),
        ]
        koji_label_font = Font(name="ＭＳ Ｐゴシック", size=10, bold=True)
        koji_value_font = Font(name="ＭＳ Ｐゴシック", size=10)
        for lbl, val in koji_info_lines:
            assert_clean(lbl, "koji label")
            assert_clean(str(val), "koji value")
            _set(ws, r, 1, lbl, font=koji_label_font, border=True,
                 align=Alignment(horizontal="center", vertical="center"))
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            _set(ws, r, 2, str(val), font=koji_value_font, border=True,
                 align=Alignment(horizontal="left", vertical="center", wrap_text=True))
            r += 1
        # 注釈（簡潔に）
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        _set(ws, r, 1,
             "※ 査定地と類似性の高い公示地です。",
             font=Font(name="ＭＳ Ｐゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 20
        r += 2

    # ■ 査定価格 / 参考価格
    section_label = "■ 参考価格" if is_degraded else "■ 机上査定価格"
    _section_header(ws, r, section_label, end_col=6)
    r += 1
    if is_degraded:
        text = (f"上記を踏まえた本物件の参考価格は {_format_price_full(rng['central_total'], target_area)} です。"
                f"（取引事例件数が不足しているため、正式な机上査定書ではありません）")
    else:
        text = f"上記を踏まえた本物件の机上査定価格は {_format_price_full(rng['central_total'], target_area)} となります。"
    assert_clean(text, "final")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws, r, 1, text, font=BIG_VALUE_FONT, align=Alignment(wrap_text=True, vertical="center"))
    ws.row_dimensions[r].height = 50
    r += 1

    # 価格直下の短縮版ディスクレーマー（読まないユーザー対策・必須UX）
    short_disc = (
        f"※ 上記は {asof.isoformat()} 時点の机上査定（参考値）です。"
        "現地・役所・法務局調査未実施。成約価格・担保評価額・鑑定評価額とは一致しません。"
        "第三者提示、金融機関提出、訴訟・税務用途には使用できません。"
    )
    assert_clean(short_disc, "short disclaimer")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws, r, 1, short_disc,
         font=Font(name="游ゴシック", size=9, italic=True, color="C00000"),
         fill=PatternFill("solid", fgColor="FFF2CC"),
         align=Alignment(wrap_text=True, vertical="center"))
    ws.row_dimensions[r].height = 38
    r += 2

    # ■ 比準表（取引事例比較表による試算）— 松田テンプレート準拠
    if hijun_rows:
        primary_h = next((h for h in hijun_rows if h.get("順位") == "規範性の高い事例"), hijun_rows[0])
        _section_header(ws, r, "■ 比準表（取引事例比較表による試算）", end_col=6)
        r += 1
        # テンプレートのフォント・書式
        TMPL_FONT = Font(name="ＭＳ Ｐゴシック", size=11)
        TMPL_FONT_BOLD = Font(name="ＭＳ Ｐゴシック", size=11, bold=True)
        TMPL_FONT_SHISAN = Font(name="ＭＳ Ｐゴシック", size=11, bold=True, color="C00000")
        ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
        ALIGN_CENTER_H = Alignment(horizontal="center")
        NUM_FMT = "#,##0_);[Red](#,##0)"
        BORDER_FULL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        BORDER_TOP = Border(left=THIN, right=THIN, top=THIN)
        BORDER_BOT = Border(left=THIN, right=THIN, bottom=THIN)

        # 事例番号（1行、B-C 列）
        _set(ws, r, 2, "事例番号", font=TMPL_FONT_BOLD, border=BORDER_FULL, align=ALIGN_CENTER_H)
        _set(ws, r, 3, primary_h.get("事例番号", "?"),
             font=TMPL_FONT, border=BORDER_FULL, align=ALIGN_CENTER_H)
        r += 1

        # 取引価格（1行）
        _set(ws, r, 2, "取引価格（円/㎡）", font=TMPL_FONT_BOLD, border=BORDER_FULL, align=ALIGN_CENTER_H)
        price_cell = ws.cell(row=r, column=3, value=int(primary_h["取引価格"]))
        price_cell.font = TMPL_FONT
        price_cell.border = BORDER_FULL
        price_cell.alignment = ALIGN_CENTER_H
        price_cell.number_format = NUM_FMT
        price_row = r
        r += 1

        # 事情補正（2行、B列縦マージ、C列に分子/分母）
        ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        _set(ws, r, 2, "事情補正", font=TMPL_FONT_BOLD, border=BORDER_FULL, align=ALIGN_CENTER)
        jijo_apply = primary_h.get("事情補正_適用", False)
        jijo_num_val = round(primary_h["事情補正"] * 100, 1)
        if jijo_num_val == int(jijo_num_val):
            jijo_num_val = int(jijo_num_val)
        _set(ws, r, 3, jijo_num_val, font=TMPL_FONT, border=BORDER_TOP, align=ALIGN_CENTER_H)
        jijo_num_row = r
        r += 1
        _set(ws, r, 3, "―" if not jijo_apply else 100,
             font=TMPL_FONT, border=BORDER_BOT, align=ALIGN_CENTER_H)
        r += 1

        # 時点修正（2行）
        ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        _set(ws, r, 2, "時点修正", font=TMPL_FONT_BOLD, border=BORDER_FULL, align=ALIGN_CENTER)
        time_num_val = round(primary_h["時点修正"] * 100, 1)
        if time_num_val == int(time_num_val):
            time_num_val = int(time_num_val)
        _set(ws, r, 3, time_num_val, font=TMPL_FONT, border=BORDER_TOP, align=ALIGN_CENTER_H)
        time_num_row = r
        r += 1
        _set(ws, r, 3, 100, font=TMPL_FONT, border=BORDER_BOT, align=ALIGN_CENTER_H)
        r += 1

        # 形状補正（2行、100/分母 形式）
        ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        _set(ws, r, 2, "形状補正", font=TMPL_FONT_BOLD, border=BORDER_FULL, align=ALIGN_CENTER)
        shape_mult = float(primary_h["標準化補正"])
        # 案件評点（=mult*100）を分母に表示。倍率は 下/上 で計算
        shape_den_val = round(shape_mult * 100, 1) if shape_mult > 0 else 100
        if shape_den_val == int(shape_den_val):
            shape_den_val = int(shape_den_val)
        _set(ws, r, 3, 100, font=TMPL_FONT, border=BORDER_TOP, align=ALIGN_CENTER_H)
        shape_num_row = r
        r += 1
        _set(ws, r, 3, shape_den_val, font=TMPL_FONT, border=BORDER_BOT, align=ALIGN_CENTER_H)
        shape_den_row = r
        r += 1

        # 地域格差（2行、100/分母 形式）
        ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        _set(ws, r, 2, "地域格差", font=TMPL_FONT_BOLD, border=BORDER_FULL, align=ALIGN_CENTER)
        chi_mult = float(primary_h["地域格差"])
        # 相乗積と一致する案件評点（=mult*100）を分母に
        chi_den_val = round(chi_mult * 100, 1) if chi_mult > 0 else 100
        if chi_den_val == int(chi_den_val):
            chi_den_val = int(chi_den_val)
        _set(ws, r, 3, 100, font=TMPL_FONT, border=BORDER_TOP, align=ALIGN_CENTER_H)
        chi_num_row = r
        r += 1
        _set(ws, r, 3, chi_den_val, font=TMPL_FONT, border=BORDER_BOT, align=ALIGN_CENTER_H)
        chi_den_row = r
        r += 1

        # 標準画地の試算値（Excel関数式、2行マージ）
        ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        ws.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=3)
        _set(ws, r, 2, "標準画地の試算値",
             font=TMPL_FONT_BOLD, border=BORDER_FULL, align=ALIGN_CENTER)
        # 標準化補正・地域格差は「上=100, 下=案件評点」配置 → 倍率 = 下/上
        expr = (f"C{price_row}*C{jijo_num_row}/100*C{time_num_row}/100"
                f"*C{shape_den_row}/C{shape_num_row}*C{chi_den_row}/C{chi_num_row}")
        formula = f"=ROUND({expr},-(LEN(INT({expr}))-3))"
        formula_cell = ws.cell(row=r, column=3, value=formula)
        formula_cell.font = TMPL_FONT_SHISAN
        formula_cell.border = BORDER_FULL
        formula_cell.alignment = ALIGN_CENTER
        formula_cell.number_format = NUM_FMT
        shisan_row = r  # 標準画地の試算値 行（査定価格formula参照用）
        r += 2

        # ■ 個別格差（角地・方位・不整形の本物件固有の格差を反映）
        # 添付参考のように、ラベルと数値を「青字」で表示して標準化補正と差別化
        BLUE_LABEL_FONT = Font(name="ＭＳ Ｐゴシック", size=11, color="2F5496")
        BLUE_VALUE_FONT = Font(name="ＭＳ Ｐゴシック", size=11, color="2F5496")
        SECTION_BLUE_FONT = Font(name="ＭＳ Ｐゴシック", size=11, bold=True, color="FFFFFF")
        SECTION_BLUE_FILL = PatternFill("solid", fgColor="2F5496")

        # 標準画地の試算値の下に縦並びで表示（B-C 列にセクションヘッダ）
        _set(ws, r, 2, "■ 個別格差",
             font=SECTION_BLUE_FONT, fill=SECTION_BLUE_FILL,
             border=BORDER_FULL, align=ALIGN_CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 1

        def _fmt_kobetsu(v):
            """個別格差値の表示用整形（整数化、±0は0）"""
            iv = round(v, 1)
            return int(iv) if iv == int(iv) else iv

        # v1.2.1: target が中間画地（角地補正率(%)未入力 or 0）の場合は角地行を非表示
        target_kado_val = primary_h.get("個別格差_角地", 0)
        kado_row = None
        if abs(round(target_kado_val, 1)) >= 0.05:
            _set(ws, r, 2, "角地（角地）", font=BLUE_LABEL_FONT, border=BORDER_FULL, align=ALIGN_CENTER_H)
            _set(ws, r, 3, _fmt_kobetsu(target_kado_val),
                 font=BLUE_VALUE_FONT, border=BORDER_FULL, align=ALIGN_CENTER_H)
            kado_row = r
            r += 1

        # 方位（青字、target の方位をラベルに）
        target_dir = str(target.get("前面道路:方位", "")).strip()
        houi_label = f"方位（{target_dir}）" if target_dir else "方位"
        _set(ws, r, 2, houi_label, font=BLUE_LABEL_FONT, border=BORDER_FULL, align=ALIGN_CENTER_H)
        houi_val = _fmt_kobetsu(primary_h.get("個別格差_方位", 0))
        _set(ws, r, 3, houi_val, font=BLUE_VALUE_FONT, border=BORDER_FULL, align=ALIGN_CENTER_H)
        houi_row = r
        r += 1

        # 不整形（青字、v1.2.1: target の土地形状をラベルに）
        target_shape = str(target.get("土地の形状", "")).strip()
        fusei_label = f"不整形（{target_shape}）" if target_shape else "不整形"
        _set(ws, r, 2, fusei_label, font=BLUE_LABEL_FONT, border=BORDER_FULL, align=ALIGN_CENTER_H)
        fusei_val = _fmt_kobetsu(primary_h.get("個別格差_不整形", 0))
        _set(ws, r, 3, fusei_val, font=BLUE_VALUE_FONT, border=BORDER_FULL, align=ALIGN_CENTER_H)
        fusei_row = r
        r += 1

        # 総和（Excel関数式、黒字）— 表示中の格差行のみを積算
        _set(ws, r, 2, "総和", font=TMPL_FONT_BOLD, border=BORDER_FULL, align=ALIGN_CENTER_H)
        factor_refs_k = []
        if kado_row is not None:
            factor_refs_k.append(f"(100+C{kado_row})/100")
        factor_refs_k.append(f"(100+C{houi_row})/100")
        factor_refs_k.append(f"(100+C{fusei_row})/100")
        soan_formula = "=" + "*".join(factor_refs_k) + "*100"
        soan_cell = ws.cell(row=r, column=3, value=soan_formula)
        soan_cell.font = TMPL_FONT
        soan_cell.border = BORDER_FULL
        soan_cell.alignment = ALIGN_CENTER_H
        soan_cell.number_format = "0.00"
        soan_row = r
        r += 2  # 1行空ける

        # 案件査定価格（ラベル青字、値は赤字）
        BLUE_BOLD_LABEL = Font(name="ＭＳ Ｐゴシック", size=11, bold=True, color="2F5496")
        _set(ws, r, 2, "案件査定価格（円/㎡）",
             font=BLUE_BOLD_LABEL, border=BORDER_FULL, align=ALIGN_CENTER_H)
        anken_inner_k = f"C{shisan_row}*C{soan_row}"
        anken_formula = f"=ROUND({anken_inner_k},-(LEN(INT({anken_inner_k}))-3))/100"
        anken_cell = ws.cell(row=r, column=3, value=anken_formula)
        anken_cell.font = TMPL_FONT_SHISAN  # red bold
        anken_cell.border = BORDER_FULL
        anken_cell.alignment = ALIGN_CENTER_H
        anken_cell.number_format = NUM_FMT
        r += 2

        # 注釈
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        _set(ws, r, 1,
             "※ 「事情補正」は売主・買主の事情が取引価格に影響している場合の調整、"
             "「時点修正」は取引時期と査定時点の地価変動による調整、"
             "「形状補正」「地域格差」はそれぞれ事例地と本物件の形状・地域条件の差を反映しています。"
             "「個別格差」は本物件固有の角地・方位・不整形による調整で、"
             "標準画地の試算値に乗じて案件査定価格を算出します。",
             font=Font(name="ＭＳ Ｐゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 40
        r += 2

    # ■ 査定の考え方
    _section_header(ws, r, "■ 査定の考え方", end_col=6)
    r += 1
    sentences = [
        "本査定は、本地区とその周辺で本物件と規範性の高い取引事例を複数選定し、",
        "それぞれの単価を本物件の特徴（面積・最寄駅までの距離・形状・接道など）に合わせて調整したうえで、査定価格を算出しています。",
        "なお、最も規範性の高い1件を「規範性の高い取引事例」として表示しています。",
        "標準価格や地価の動きとの整合も確認しています。",
        "あくまで一次査定であり、現地確認や市場動向によって最終価格は変動し得ます。",
    ]
    text = "".join(sentences)
    assert_clean(text, "story")
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=6)
    _set(ws, r, 1, text, font=VALUE_FONT, align=Alignment(wrap_text=True, vertical="top"))
    ws.row_dimensions[r].height = 80
    r += 4

    # ■ 重要事項（机上査定の前提と免責）
    _section_header(ws, r, "■ 重要事項（机上査定の前提と免責）", end_col=6)
    r += 1

    # 1) 自動生成・鑑定評価書ではない
    text_1 = (
        "本書はご入力いただいた情報および国土交通省「不動産取引価格情報」「地価公示」等の"
        "公開データに基づき、ソフトウェアにより自動生成した机上査定です。"
        "本査定額は自動算出によるものであり、不動産鑑定評価基準に基づく不動産鑑定評価書ではなく、"
        "個別の不動産鑑定士による判断を経たものでもありません。"
    )
    assert_clean(text_1, "important note 1")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws, r, 1, text_1, font=VALUE_FONT,
         align=Alignment(wrap_text=True, vertical="top"))
    ws.row_dimensions[r].height = 70
    r += 1

    # 2) 行っていない調査
    label_b = Font(name="游ゴシック", size=10, bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws, r, 1, "本査定では、以下の調査を行っておりません。", font=label_b)
    r += 1
    skipped_surveys = [
        "対象不動産の現地調査（外観・内部・接道・近隣環境・越境・境界等）",
        "役所調査（公法上の規制、道路種別、インフラ、開発許可、建築確認履歴等）",
        "法務局調査（登記・公図・地積測量図・権利関係の精査）",
        "賃貸借契約・修繕履歴・個別契約条件の精査",
        "土壌汚染・地下埋設物・アスベスト等の物的リスク調査",
    ]
    for s in skipped_surveys:
        assert_clean(s, "skipped survey")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        _set(ws, r, 1, f"　・{s}", font=VALUE_FONT)
        r += 1
    r += 1

    # 3) 参考価格としての位置づけ
    text_3 = (
        "本査定額は、入力情報の正確性および公開情報の精度に依存する参考価格であり、"
        "実際の成約価格、金融機関の担保評価額、税務評価額、不動産鑑定評価額とは一致しません。"
        "個別の減価要因（境界未確定、越境、再建築不可、心理的瑕疵、土壌汚染等）が存在する場合、"
        "査定額は大きく変動します。"
    )
    assert_clean(text_3, "important note 3")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws, r, 1, text_3, font=VALUE_FONT,
         align=Alignment(wrap_text=True, vertical="top"))
    ws.row_dimensions[r].height = 60
    r += 1

    # 4) 利用範囲制限
    text_4 = (
        "本書は、利用者様ご自身の検討用途にのみご利用いただくものとし、"
        "第三者への提示・交付、訴訟・調停等の証拠資料、金融機関への提出資料、税務申告等の用途には使用できません。"
        "これらの用途には、不動産鑑定士による鑑定評価書の取得を推奨いたします。"
    )
    assert_clean(text_4, "important note 4")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws, r, 1, text_4, font=VALUE_FONT,
         align=Alignment(wrap_text=True, vertical="top"))
    ws.row_dimensions[r].height = 50
    r += 1

    # 5) 損害免責
    text_5 = (
        "本書の利用に起因して利用者様または第三者に生じた損害について、"
        "本ソフトウェア提供元およびその運営者は一切の責任を負いません。"
        "詳細は別途定める利用規約をご確認ください。"
    )
    assert_clean(text_5, "important note 5")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws, r, 1, text_5, font=VALUE_FONT,
         align=Alignment(wrap_text=True, vertical="top"))
    ws.row_dimensions[r].height = 40
    r += 2

    # 価格時点を末尾で再表示（査定額と同じ視認性で）
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws, r, 1, f"価格時点：{asof.isoformat()}",
         font=Font(name="游ゴシック", size=12, bold=True),
         align=Alignment(horizontal="right"))
    r += 1

    _adjust_col_widths(ws, [12, 26, 16, 18, 14, 14])


def _apply_page_setup(wb: Workbook, target: dict):
    """各シートに印刷設定を適用。
    業者用: A3 横 + 縮小印刷（fitToWidth=1）
    グラフ: A4 横 + 縮小印刷（fitToPage）
    顧客用: A4 縦 + ヘッダ「机上査定書」+ フッタ「N / 総页」
    """
    from openpyxl.worksheet.page import PageMargins
    # PAPERSIZE: A3=8, A4=9
    A3_SIZE, A4_SIZE = 8, 9
    target_label = target.get("物件略号", "")
    location = f"{target.get('市区町村名','')} {target.get('地区名','')}{target.get('丁目','')}"

    # 業者用：A3 横、幅に合わせて縮小、印刷タイトルとして 1行目を固定
    if "業者用" in wb.sheetnames:
        ws = wb["業者用"]
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = A3_SIZE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 縦は複数ページ可
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6,
                                       header=0.3, footer=0.3)
        # 1行目を印刷タイトルとして固定（各ページにタイトルが表示される）
        ws.print_title_rows = "1:1"
        # ヘッダ・フッタ
        ws.oddHeader.left.text = f"業者用 — {target_label}"
        ws.oddHeader.right.text = "&D"  # 日付
        ws.oddFooter.center.text = "&P / &N"

    # グラフ：A4 横、ページに合わせて縮小
    if "グラフ" in wb.sheetnames:
        gs = wb["グラフ"]
        gs.page_setup.orientation = gs.ORIENTATION_LANDSCAPE
        gs.page_setup.paperSize = A4_SIZE
        gs.page_setup.fitToWidth = 1
        gs.page_setup.fitToHeight = 1
        gs.sheet_properties.pageSetUpPr.fitToPage = True
        gs.print_options.horizontalCentered = True
        gs.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.6)
        gs.oddHeader.left.text = f"附属資料 — {target_label}"
        gs.oddFooter.center.text = "&P / &N"

    # 顧客用：A4 縦、ヘッダ「机上査定書」、フッタ「現/総」
    if "顧客用" in wb.sheetnames:
        ks = wb["顧客用"]
        ks.page_setup.orientation = ks.ORIENTATION_PORTRAIT
        ks.page_setup.paperSize = A4_SIZE
        ks.page_setup.fitToWidth = 1
        ks.page_setup.fitToHeight = 0
        ks.sheet_properties.pageSetUpPr.fitToPage = True
        ks.print_options.horizontalCentered = True
        ks.page_margins = PageMargins(left=0.6, right=0.6, top=0.8, bottom=0.8,
                                       header=0.3, footer=0.4)
        ks.oddHeader.center.text = "&\"游ゴシック,Bold\"&14机上査定書"
        ks.oddHeader.right.text = location
        ks.oddFooter.center.text = "&P / &N"


def write_xlsx(ctx: dict, output_path: Path) -> Path:
    wb = Workbook()
    # デフォルトシートを削除
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _write_gyosha_sheet(wb, ctx)
    _write_kokyaku_sheet(wb, ctx)
    # 印刷設定（A3横/A4横/A4縦）
    _apply_page_setup(wb, ctx.get("target", {}))
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
